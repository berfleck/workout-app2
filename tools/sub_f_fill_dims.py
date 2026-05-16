"""
Sub-F: Preencher 5 colunas dims para os 59 exercícios non-YAML.

Regras aplicadas:
- pegada: null para todos (exceto Desenv. Landmine = 'neutra' via C3)
- plano_corporal: 'em_pe' ou 'deitado' para hinge; null para resto
- equipamento_grupo: por análise individual de eq_primario + grupo
- variante_pontual: False para todos
- variacao_de: NÃO alterado (já preenchido ou null por decisão)

Decisões C1-C7 aplicadas:
  C1: flexao_plantar → pegada=null (já satisfeito pela regra geral)
  C3: Desenv. Landmine → pegada='neutra'
  C4: anilha → halter (exceto Lev. Terra Anilha → barra, Supino Com Anilha → null)
  C5: Feijão/Slide knee_flex → equipamento_grupo=null
  C6: Hip Thrust Uni., Ponte Unilateral, Copenhagen → corporal
  C7: Hiperextensão 45° → plano_corporal='em_pe'
"""

import openpyxl
import sys

XLSX_PATH = "banco_exercicios.xlsx"

# ── Mapping: nome_exacto → {dim: valor}
# Para dims não listadas: pegada=None, plano=None, equip depende
# variante_pontual = False sempre
# 'pegada' só inclui quando != None (C3 é a única exceção)
# 'plano_corporal' só inclui quando != None (hinges)

UPDATES = {
    # ── HINGE — em pé ──────────────────────────────────────────
    "Stiff Barra Livre":    {"plano_corporal": "em_pe", "equipamento_grupo": "barra"},
    "Stiff Barra Smith":    {"plano_corporal": "em_pe", "equipamento_grupo": "barra_guiada"},
    "Stiff Halteres":       {"plano_corporal": "em_pe", "equipamento_grupo": "halter"},
    "Stiff Uni. Halteres":  {"plano_corporal": "em_pe", "equipamento_grupo": "halter"},
    "Stiff Uni. Smith":     {"plano_corporal": "em_pe", "equipamento_grupo": "barra_guiada"},
    "Good Morning":         {"plano_corporal": "em_pe", "equipamento_grupo": "barra"},
    "Hiperextensão 45°":    {"plano_corporal": "em_pe", "equipamento_grupo": "maquina"},  # C7
    "Lev. Terra":           {"plano_corporal": "em_pe", "equipamento_grupo": "barra"},
    "Lev. Terra Anilha":    {"plano_corporal": "em_pe", "equipamento_grupo": "barra"},    # C4-exc (barra, não halter)
    "Lev. Terra Sumô":      {"plano_corporal": "em_pe", "equipamento_grupo": "barra"},

    # ── HINGE — deitado ────────────────────────────────────────
    "Hip Thrust":           {"plano_corporal": "deitado", "equipamento_grupo": "barra"},
    "Hip Thrust C/ Band":   {"plano_corporal": "deitado", "equipamento_grupo": "banda_elastica"},
    "Hip Thrust Uni.":      {"plano_corporal": "deitado", "equipamento_grupo": "corporal"},   # C6
    "Ponte":                {"plano_corporal": "deitado", "equipamento_grupo": "corporal"},
    "Ponte Alternada":      {"plano_corporal": "deitado", "equipamento_grupo": "corporal"},
    "Ponte C/ Band":        {"plano_corporal": "deitado", "equipamento_grupo": "banda_elastica"},
    "Ponte Na Caixa":       {"plano_corporal": "deitado", "equipamento_grupo": "caixa"},
    "Ponte Uni. Caixa":     {"plano_corporal": "deitado", "equipamento_grupo": "caixa"},
    "Ponte Unilateral":     {"plano_corporal": "deitado", "equipamento_grupo": "corporal"},   # C6

    # ── KNEE_FLEXION ────────────────────────────────────────────
    "Cadeira Flexora":          {"equipamento_grupo": "maquina"},
    "Flexão Joelhos Feijão":    {"equipamento_grupo": None},   # C5
    "Flexão Joelhos Slide":     {"equipamento_grupo": None},   # C5
    "Nordic Curl":              {"equipamento_grupo": "corporal"},

    # ── ABDUCTION ───────────────────────────────────────────────
    "Abdução Polia":            {"equipamento_grupo": "polia"},
    "Desloc. Lateral c/ Band":  {"equipamento_grupo": "banda_elastica"},
    "Side Clams":               {"equipamento_grupo": "banda_elastica"},   # eq_primario = Glute band

    # ── ADDUCTION ───────────────────────────────────────────────
    "Adução Polia":             {"equipamento_grupo": "polia"},
    "Copenhagen Adduction":     {"equipamento_grupo": "corporal"},   # C6

    # ── FLEXAO_PLANTAR ──────────────────────────────────────────
    # C1: pegada=null (regra geral já cobre); eq_primario=Rack → corporal (bodyweight + rack para equilíbrio)
    "Elevação De Panturrilha Em Pé":      {"equipamento_grupo": "corporal"},
    "Elevação Unilateral Panturrilha":    {"equipamento_grupo": "corporal"},

    # ── OMBRO_COMPOSTO ──────────────────────────────────────────
    "Desenv. Halteres Sentado": {"equipamento_grupo": "halter"},
    "Desenv. Halteres Uni.":    {"equipamento_grupo": "halter"},
    "Desenv. Landmine":         {"pegada": "neutra", "equipamento_grupo": "barra"},  # C3
    "Desenvolvimento Barra":    {"equipamento_grupo": "barra"},   # eq_primario=Rack (barra no rack)
    "Desenvolvimento Smith":    {"equipamento_grupo": "barra_guiada"},

    # ── OMBRO_ISOLADO ───────────────────────────────────────────
    "Elevação Frontal Anilha":  {"equipamento_grupo": "halter"},   # C4: anilha→halter
    "Elevação Frontal Halteres":{"equipamento_grupo": "halter"},
    "Elevação Lateral":         {"equipamento_grupo": "halter"},
    "Elevação Lateral Polia":   {"equipamento_grupo": "polia"},
    "Elevação Lateral Sentado": {"equipamento_grupo": "halter"},

    # ── POSTERIOR_OMBRO ─────────────────────────────────────────
    "Crucifíxo Invertido":  {"equipamento_grupo": "halter"},
    "Face Pull (Polia)":        {"equipamento_grupo": "polia"},    # C2: pegada=null (já cobre regra geral)
    "Posterior Ombro Polia":    {"equipamento_grupo": "polia"},

    # ── BICEPS ──────────────────────────────────────────────────
    "Bíceps 21S":       {"equipamento_grupo": "barra"},
    "Bíceps Banco":     {"equipamento_grupo": "halter"},
    "Bíceps Bayesian":  {"equipamento_grupo": "polia"},
    "Bíceps Cabo":      {"equipamento_grupo": "polia"},
    "Bíceps Halteres":  {"equipamento_grupo": "halter"},
    "Bíceps Martelo":   {"equipamento_grupo": "halter"},

    # ── TRICEPS ─────────────────────────────────────────────────
    # G7 spec: pegada=null (regra geral já cobre)
    "Tríceps Coice Com Halter": {"equipamento_grupo": "halter"},
    "Tríceps Coice Polia":      {"equipamento_grupo": "polia"},
    "Tríceps Corda":            {"equipamento_grupo": "polia"},
    "Tríceps Francês":          {"equipamento_grupo": "halter"},   # eq_primario=Halteres
    "Tríceps Mergulho Banco":   {"equipamento_grupo": "corporal"}, # bench dips = corporal
    "Tríceps Polia Alta":       {"equipamento_grupo": "polia"},
    "Tríceps Testa Halteres":   {"equipamento_grupo": "halter"},
    "Tríceps Unilateral Polia": {"equipamento_grupo": "polia"},

    # ── CARDIO ──────────────────────────────────────────────────
    "Air Bike (Sprint)":       {},  # all dims null; só variante_pontual=False
    "Air Bike (Steady State)": {},
}

# Verify count
assert len(UPDATES) == 59, f"Expected 59 entries, got {len(UPDATES)}"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    COL = {h: i + 1 for i, h in enumerate(headers) if h}

    col_pegada    = COL["pegada"]
    col_plano     = COL["plano_corporal"]
    col_equip     = COL["equipamento_grupo"]
    col_vp        = COL["variante_pontual"]

    found = set()
    updated_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        nome_cell = row[COL["nome"] - 1]
        if nome_cell.value is None:
            continue
        nome = str(nome_cell.value).strip()

        if nome not in UPDATES:
            continue

        overrides = UPDATES[nome]

        # pegada: None by default, 'neutra' for Landmine (C3)
        pegada_val = overrides.get("pegada", None)
        # plano_corporal: None by default, em_pe/deitado for hinge
        plano_val = overrides.get("plano_corporal", None)
        # equipamento_grupo: per override (None = clear cell)
        equip_val = overrides.get("equipamento_grupo", "NOT_SET")

        ws.cell(row=row_idx, column=col_pegada).value = pegada_val
        ws.cell(row=row_idx, column=col_plano).value  = plano_val
        if equip_val != "NOT_SET":
            ws.cell(row=row_idx, column=col_equip).value = equip_val
        else:
            ws.cell(row=row_idx, column=col_equip).value = None
        ws.cell(row=row_idx, column=col_vp).value = False

        found.add(nome)
        updated_rows.append((row_idx, nome, pegada_val, plano_val,
                              equip_val if equip_val != "NOT_SET" else None))

    # Report
    not_found = set(UPDATES.keys()) - found
    print(f"Updated: {len(found)}/59 exercises")
    if not_found:
        print(f"\nNOT FOUND in XLSX ({len(not_found)}):")
        for n in sorted(not_found):
            print(f"  - {repr(n)}")
        sys.exit(1)

    print("\nSample updates:")
    for row_idx, nome, p, pl, eq in updated_rows[:10]:
        print(f"  row {row_idx}: {nome} | pegada={p} | plano={pl} | equip={eq}")

    wb.save(XLSX_PATH)
    print(f"\nSaved {XLSX_PATH}")


if __name__ == "__main__":
    main()
