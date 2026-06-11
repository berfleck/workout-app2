# -*- coding: utf-8 -*-
"""Gera planilha de curadoria mapeando exercicios do banco (PT) -> imagens da
free-exercise-db (EN, dominio publico).

Saida: mapeamento_midia.xlsx na raiz do projeto, com 3 candidatos por exercicio,
cada um com link clicavel pra foto. O usuario (PT) revisa e marca a escolha.

Re-rodar e seguro: NAO sobrescreve escolhas ja feitas se a planilha existir
(faz merge pela coluna 'nome' + 'escolha').
"""
import json
import os
import sys
import unicodedata
import urllib.request
from difflib import SequenceMatcher

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BANCO = os.path.join(RAIZ, "banco_exercicios.xlsx")
SAIDA = os.path.join(RAIZ, "mapeamento_midia.xlsx")
FEDB_URL = "https://raw.githubusercontent.com/yuhonas/free-exercise-db/main/dist/exercises.json"
IMG_BASE = "https://raw.githubusercontent.com/yuhonas/free-exercise-db/main/exercises/"

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Curadoria de mídia — BF Treinamento</title>
<style>
  :root { --laranja:#e85d04; --bg:#f9fafb; }
  * { box-sizing:border-box; }
  body { font-family:"DM Sans",system-ui,Arial,sans-serif; margin:0; background:var(--bg); color:#1f2937; }
  header { position:sticky; top:0; z-index:10; background:#fff; border-bottom:2px solid var(--laranja);
           padding:12px 20px; display:flex; align-items:center; gap:16px; flex-wrap:wrap; box-shadow:0 1px 4px rgba(0,0,0,.06); }
  header h1 { font-size:18px; margin:0; color:var(--laranja); }
  .prog { font-weight:600; }
  .prog b { color:var(--laranja); }
  button, select { font:inherit; padding:8px 14px; border-radius:8px; border:1px solid #d1d5db; background:#fff; cursor:pointer; }
  button.primary { background:var(--laranja); color:#fff; border-color:var(--laranja); font-weight:600; }
  main { max-width:1100px; margin:0 auto; padding:20px; }
  .ex { background:#fff; border-radius:12px; padding:16px; margin-bottom:18px; box-shadow:0 1px 3px rgba(0,0,0,.08); }
  .ex.decidido { opacity:.6; }
  .ex-head { display:flex; align-items:baseline; gap:10px; flex-wrap:wrap; margin-bottom:12px; }
  .ex-head h2 { font-size:17px; margin:0; }
  .meta { font-size:13px; color:#6b7280; }
  .badge { font-size:11px; font-weight:700; padding:2px 8px; border-radius:99px; color:#fff; }
  .badge.Forte { background:#16a34a; } .badge.Media { background:#d97706; } .badge.Fraca { background:#dc2626; }
  .cands { display:grid; grid-template-columns:repeat(auto-fill,minmax(180px,1fr)); gap:12px; }
  .cand { border:3px solid transparent; border-radius:10px; padding:8px; cursor:pointer; text-align:center; background:#fafafa; transition:.12s; }
  .cand:hover { background:#f3f4f6; }
  .cand.sel { border-color:var(--laranja); background:#fff7ed; }
  .cand img { width:100%; height:160px; object-fit:contain; background:#fff; border-radius:6px; }
  .cand .nm { font-size:12px; margin-top:6px; line-height:1.3; }
  .cand.nenhum, .cand.buscar { display:flex; align-items:center; justify-content:center; min-height:200px; font-weight:600; color:#6b7280; text-align:center; }
  .cand.nenhum.sel { color:var(--laranja); }
  .cand.buscar { border:2px dashed #d1d5db; }
  .cand.buscar:hover { border-color:var(--laranja); color:var(--laranja); }
  .busca-box { margin-top:12px; padding:12px; background:#f3f4f6; border-radius:10px; }
  .busca-box input { width:100%; padding:10px; border:1px solid #d1d5db; border-radius:8px; font:inherit; }
  .busca-res { display:grid; grid-template-columns:repeat(auto-fill,minmax(150px,1fr)); gap:10px; margin-top:10px; max-height:420px; overflow:auto; }
  .busca-res .cand img { height:120px; }
  .filtros { margin-left:auto; display:flex; gap:8px; align-items:center; }
  .hint { font-size:12px; color:#6b7280; width:100%; }
</style>
</head>
<body>
<header>
  <h1>🖼️ Curadoria de mídia</h1>
  <span class="prog">Decididos: <b id="cnt">0</b> / <span id="tot">0</span></span>
  <div class="filtros">
    <select id="filtro">
      <option value="todos">Mostrar todos</option>
      <option value="pendentes">Só pendentes</option>
      <option value="fraca">Só vermelhos (Fraca)</option>
    </select>
    <button class="primary" id="baixar">⬇ Baixar escolhas</button>
  </div>
  <div class="hint">As fotos <b>alternam início↔fim</b> sozinhas (simula o movimento). Clique no candidato certo, em <b>🔍 Buscar outro</b> pra procurar qualquer um dos 873 (em inglês), ou em "Nenhum serve". Salva automático — ao terminar, clique <b>Baixar escolhas</b> e me mande o arquivo.</div>
</header>
<main id="lista"></main>

<script>
const DADOS = __DADOS__;
const TODOS = __TODOS__;
const PORID = {}; TODOS.forEach(e => PORID[e.id] = e);
const LS = "curadoria_midia_v1";
let escolhas = JSON.parse(localStorage.getItem(LS) || "{}");
let buscaAberta = new Set();

function salvar(){ localStorage.setItem(LS, JSON.stringify(escolhas)); render(); }

function escolher(nome, val, media_id){
  escolhas[nome] = {escolha: val, media_id: media_id || ""};
  salvar();
}
function escolherManual(nome, id){
  const e = PORID[id];
  escolhas[nome] = {escolha:"manual", media_id:id, media_name: e ? e.name : id};
  buscaAberta.delete(nome);
  salvar();
}
function toggleBusca(nome){
  if (buscaAberta.has(nome)) buscaAberta.delete(nome); else buscaAberta.add(nome);
  render();
}

function tileHTML(nome, c, val, sel){
  return `<div class="cand${sel?' sel':''}" onclick="${val==='manual'
      ? `escolherManual('${esc(nome)}','${esc(c.id)}')`
      : `escolher('${esc(nome)}','${val}','${esc(c.id)}')`}">
    <img loading="lazy" data-a="${c.img0}" data-b="${c.img1}" src="${c.img0}" alt="">
    <div class="nm">${c.name}</div></div>`;
}

function buscar(i, nome, q){
  const out = document.getElementById("res"+i);
  q = q.trim().toLowerCase();
  if (q.length < 2){ out.innerHTML = '<span class="hint">Digite ao menos 2 letras (em inglês: squat, row, curl...)</span>'; return; }
  const res = TODOS.filter(e => e.name.toLowerCase().includes(q)).slice(0, 30);
  out.innerHTML = res.length
    ? res.map(c => tileHTML(nome, c, "manual", false)).join("")
    : '<span class="hint">Nada encontrado.</span>';
}

function render(){
  const lista = document.getElementById("lista");
  const filtro = document.getElementById("filtro").value;
  lista.innerHTML = "";
  let decididos = 0;
  DADOS.forEach((ex, i) => {
    const dec = escolhas[ex.nome] !== undefined;
    if (dec) decididos++;
    if (filtro === "pendentes" && dec) return;
    if (filtro === "fraca" && ex.conf !== "Fraca") return;

    const card = document.createElement("div");
    card.className = "ex" + (dec ? " decidido" : "");
    const cur = escolhas[ex.nome] || {};
    const candIds = ex.cands.map(c => c.id);

    let tiles = ex.cands.map((c, j) =>
      tileHTML(ex.nome, c, String(j+1), cur.escolha === String(j+1))).join("");
    // pick manual fora dos 3 candidatos -> tile extra destacado
    if (cur.escolha === "manual" && !candIds.includes(cur.media_id) && PORID[cur.media_id]){
      tiles += tileHTML(ex.nome, PORID[cur.media_id], "manual", true);
    }
    tiles += `<div class="cand nenhum${cur.escolha==='nenhum'?' sel':''}" onclick="escolher('${esc(ex.nome)}','nenhum','')">✕ Nenhum serve</div>`;
    tiles += `<div class="cand buscar" onclick="toggleBusca('${esc(ex.nome)}')">🔍 Buscar outro<br>na base</div>`;

    let busca = "";
    if (buscaAberta.has(ex.nome)){
      busca = `<div class="busca-box">
        <input type="text" placeholder="Buscar em inglês: squat, bench, deadlift, row..." oninput="buscar(${i},'${esc(ex.nome)}',this.value)">
        <div class="busca-res" id="res${i}"><span class="hint">Digite ao menos 2 letras (em inglês).</span></div>
      </div>`;
    }

    card.innerHTML = `<div class="ex-head">
        <h2>${ex.nome}</h2>
        <span class="badge ${ex.conf}">${ex.conf}</span>
        <span class="meta">${ex.padrao} · ${ex.musc} · ${ex.eq||"—"}</span>
      </div><div class="cands">${tiles}</div>${busca}`;
    lista.appendChild(card);
  });
  document.getElementById("cnt").textContent = decididos;
  document.getElementById("tot").textContent = DADOS.length;
}
function esc(s){ return String(s).replace(/\\/g,"\\\\").replace(/'/g,"\\'"); }

// animacao: alterna inicio<->fim em todas as imagens visiveis
let frame = 0;
setInterval(() => {
  frame = 1 - frame;
  document.querySelectorAll(".cand img").forEach(img => {
    const t = frame ? img.dataset.b : img.dataset.a;
    if (t && img.src !== t) img.src = t;
  });
}, 900);

document.getElementById("filtro").onchange = render;
document.getElementById("baixar").onclick = () => {
  const blob = new Blob([JSON.stringify(escolhas, null, 2)], {type:"application/json"});
  const a = document.createElement("a");
  a.href = URL.createObjectURL(blob);
  a.download = "escolhas_midia.json";
  a.click();
};
render();
</script>
</body>
</html>
"""


def strip(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return s.lower()


# PT -> EN: termos de movimento e equipamento
TR = {
    "agachamento": "squat", "agach": "squat", "supino": "bench press",
    "crucifixo": "fly", "crossover": "cable crossover", "apoio": "push up",
    "flexao": "flexion", "remada": "row", "puxada": "pulldown",
    "barra fixa": "pull up", "barra": "pull up", "biceps": "biceps curl",
    "rosca": "curl", "triceps": "triceps", "martelo": "hammer curl",
    "frances": "lying triceps", "testa": "lying triceps extension",
    "coice": "kickback", "corda": "rope", "desenvolvimento": "shoulder press",
    "desenv": "shoulder press", "elevacao lateral": "lateral raise",
    "elevacao frontal": "front raise", "elevacao": "raise",
    "panturrilha": "calf raise", "face pull": "face pull",
    "posterior": "rear delt", "stiff": "stiff leg deadlift",
    "lev terra": "deadlift", "terra": "deadlift", "hip thrust": "hip thrust",
    "ponte": "glute bridge", "good morning": "good morning",
    "hiperextensao": "hyperextension", "cadeira extensora": "leg extension",
    "extensora": "leg extension", "cadeira flexora": "leg curl",
    "flexora": "leg curl", "leg press": "leg press", "passada": "lunge",
    "recuo": "lunge", "step up": "step up", "bulgaro": "bulgarian split squat",
    "walking lunges": "walking lunge", "box jump": "box jump",
    "goblet": "goblet squat", "prancha": "plank", "abdominal": "crunch",
    "crunch": "crunch", "infra": "leg raise", "hollow": "hollow",
    "dead bug": "dead bug", "russian twist": "russian twist",
    "pallof": "pallof press", "v-up": "v-up", "canoinha": "v-up",
    "roda abdominal": "ab wheel", "pulldown": "pulldown",
    "pullover": "pullover", "serrote": "one arm row", "adducao": "adduction",
    "abducao": "abduction", "aducao": "adduction", "copenhagen": "copenhagen",
    "nordic": "nordic", "air bike": "bike", "kickback": "kickback",
    "glute band": "band", "mergulho": "dip", "fechado": "close grip",
    "inclinado": "incline", "sentado": "seated", "anilha": "plate",
    "halteres": "dumbbell", "halter": "dumbbell", "polia": "cable",
    "cabo": "cable", "smith": "smith", "unilateral": "one arm",
    "uni": "one arm", "aberta": "wide grip", "supinada": "underhand",
    "neutra": "neutral grip",
}
EQMAP = {
    "halter": "dumbbell", "halteres": "dumbbell", "barra": "barbell",
    "polia": "cable", "crossover": "cable", "smith": "machine",
    "rack": "barbell", "olimpica": "barbell", "landmine": "barbell",
    "colchonete": "body only", "cadeira": "machine", "leg press": "machine",
    "banco": "body only", "glute band": "bands", "bola": "exercise ball",
    "feijao": "exercise ball", "ab wheel": "other", "caixa": "body only",
    "air bike": "machine",
}
MUSC = {
    "peitoral": "chest", "triceps": "triceps", "biceps": "biceps",
    "dorsal": "lats", "romboides": "middle back", "quadriceps": "quadriceps",
    "gluteos": "glutes", "gluteo": "glutes", "isquiotibiais": "hamstrings",
    "adutores": "adductors", "deltoide": "shoulders", "core": "abdominals",
    "abdominal": "abdominals", "obliquos": "abdominals", "gastrocnemio": "calves",
    "eretores": "lower back", "iliopsoas": "abdominals", "manguito": "shoulders",
    "serratil": "chest", "braquiorradial": "forearms",
}


def translate(nome):
    s = strip(nome)
    out = []
    for pt in sorted(TR, key=len, reverse=True):
        if pt in s:
            out.append(TR[pt])
            s = s.replace(pt, " ")
    return " ".join(out) + " " + " ".join(w for w in s.split() if len(w) > 2)


def musc_en(m):
    s = strip(m)
    return set(MUSC[k] for k in MUSC if k in s)


def eq_en(e):
    s = strip(e or "")
    for k in EQMAP:
        if k in s:
            return EQMAP[k]
    return None


def score(q_en, q_musc, q_eq, e):
    nm = SequenceMatcher(None, q_en, e["_n"]).ratio()
    toks_q = set(q_en.split())
    toks_e = set(e["_n"].split())
    overlap = len(toks_q & toks_e) / max(1, len(toks_q))
    em = e.get("primaryMuscles") or []
    mm = 1.0 if (q_musc & set(em)) else 0.0
    eqm = 1.0 if (q_eq and e.get("equipment") == q_eq) else 0.0
    return 0.45 * overlap + 0.25 * nm + 0.20 * mm + 0.10 * eqm


def confianca(s):
    if s >= 0.55:
        return "Forte"
    if s >= 0.40:
        return "Media"
    return "Fraca"


def carregar_fedb():
    cache = os.path.join(os.environ.get("TEMP", RAIZ), "fedb.json")
    if not os.path.exists(cache):
        print("Baixando free-exercise-db...")
        urllib.request.urlretrieve(FEDB_URL, cache)
    data = json.load(open(cache, encoding="utf-8"))
    for e in data:
        e["_n"] = strip(e["name"])
    return data


def carregar_banco():
    wb = openpyxl.load_workbook(BANCO, read_only=True)
    ws = wb["Exercícios"]
    rows = list(ws.iter_rows(values_only=True))
    h = list(rows[0])
    ci = {c: h.index(c) for c in ["nome", "eq_primario", "musculo_primario", "padrao"]}
    out = []
    for r in rows[1:]:
        if not r[ci["nome"]]:
            continue
        out.append({
            "nome": r[ci["nome"]],
            "eq": r[ci["eq_primario"]],
            "musc": r[ci["musculo_primario"]],
            "padrao": r[ci["padrao"]],
        })
    return out


def escolhas_existentes():
    """Le escolhas ja feitas numa planilha anterior, pra nao perder curadoria."""
    if not os.path.exists(SAIDA):
        return {}
    wb = openpyxl.load_workbook(SAIDA)
    ws = wb.active
    h = [c.value for c in ws[1]]
    try:
        i_nome = h.index("nome")
        i_esc = h.index("escolha")
        i_mid = h.index("media_id_final")
    except ValueError:
        return {}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[i_nome]:
            out[row[i_nome]] = {"escolha": row[i_esc], "media_id_final": row[i_mid]}
    return out


def computar_candidatos(banco, data):
    """Para cada exercicio do banco, retorna os 3 melhores candidatos da free-db
    com ambas as imagens (0.jpg = inicio, 1.jpg = fim)."""
    out = []
    for ex in banco:
        q_en = translate(ex["nome"])
        q_musc = musc_en(ex["musc"])
        q_eq = eq_en(ex["eq"])
        ranked = sorted(data, key=lambda e: score(q_en, q_musc, q_eq, e), reverse=True)
        top = ranked[:3]
        conf = confianca(score(q_en, q_musc, q_eq, top[0]))
        cands = []
        for e in top:
            imgs = [IMG_BASE + img for img in (e.get("images") or [])]
            cands.append({
                "id": e["id"],
                "name": e["name"],
                "img0": imgs[0] if imgs else "",
                "img1": imgs[1] if len(imgs) > 1 else (imgs[0] if imgs else ""),
            })
        out.append({
            "nome": ex["nome"], "padrao": ex["padrao"],
            "musc": ex["musc"], "eq": ex["eq"],
            "conf": conf, "cands": cands,
        })
    return out


def gerar_html(rows, data):
    """Pagina de curadoria: cada candidato anima as 2 fotos (inicio<->fim),
    usuario escolhe clicando ou busca manualmente em toda a base; exporta JSON."""
    todos = []
    for e in data:
        imgs = [IMG_BASE + img for img in (e.get("images") or [])]
        todos.append({
            "id": e["id"], "name": e["name"],
            "img0": imgs[0] if imgs else "",
            "img1": imgs[1] if len(imgs) > 1 else (imgs[0] if imgs else ""),
        })
    saida_html = os.path.join(RAIZ, "mapeamento_midia.html")
    html = (HTML_TEMPLATE
            .replace("__DADOS__", json.dumps(rows, ensure_ascii=False))
            .replace("__TODOS__", json.dumps(todos, ensure_ascii=False)))
    with open(saida_html, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Gerado: {saida_html}")


def main():
    data = carregar_fedb()
    banco = carregar_banco()
    prev = escolhas_existentes()
    print(f"{len(banco)} exercicios no banco, {len(data)} na free-exercise-db.")

    rows = computar_candidatos(banco, data)
    gerar_html(rows, data)

    wb = openpyxl.Workbook()

    # --- Aba Instrucoes ---
    inst = wb.active
    inst.title = "Instruções"
    linhas = [
        ("Como usar esta planilha", ""),
        ("", ""),
        ("1.", "Cada linha e um exercicio seu. As colunas Candidato 1/2/3 sao as melhores correspondencias de imagem."),
        ("2.", "Clique no nome de um candidato (azul, sublinhado) pra ABRIR A FOTO no navegador e ver se bate."),
        ("3.", "Na coluna 'escolha', deixe 1, 2 ou 3 pro candidato certo. Se nenhum servir, escreva 'nenhum'."),
        ("4.", "Criterio combinado: APROXIMADO SERVE. Se mostra o movimento, mesmo com equipamento/angulo um pouco diferente, vale."),
        ("5.", "Cores da coluna 'confianca': VERDE=provavel certo (so confira), AMARELO=confira com atencao, VERMELHO=provavel sem match."),
        ("6.", "Quando terminar, salve o arquivo. Eu leio a coluna 'escolha' e gero o mapa final pro app."),
        ("", ""),
        ("Obs:", "Pode parar no meio e continuar depois. Re-rodar o gerador NAO apaga suas escolhas ja feitas."),
    ]
    for r, (a, b) in enumerate(linhas, 1):
        inst.cell(r, 1, a).font = Font(bold=(r == 1), size=14 if r == 1 else 11)
        inst.cell(r, 2, b)
    inst.column_dimensions["A"].width = 6
    inst.column_dimensions["B"].width = 110

    # --- Aba Mapeamento ---
    ws = wb.create_sheet("Mapeamento")
    headers = ["nome", "padrao", "musculo", "equip", "confianca",
               "candidato 1", "candidato 2", "candidato 3",
               "escolha", "media_id_final",
               "id_1", "id_2", "id_3"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E85D04")
        c.font = Font(bold=True, color="FFFFFF")
    ws.freeze_panes = "A2"

    fill_cor = {
        "Forte": PatternFill("solid", fgColor="C6EFCE"),
        "Media": PatternFill("solid", fgColor="FFEB9C"),
        "Fraca": PatternFill("solid", fgColor="FFC7CE"),
    }
    link_font = Font(color="0563C1", underline="single")

    for ex in banco:
        q_en = translate(ex["nome"])
        q_musc = musc_en(ex["musc"])
        q_eq = eq_en(ex["eq"])
        ranked = sorted(data, key=lambda e: score(q_en, q_musc, q_eq, e), reverse=True)
        top = ranked[:3]
        conf = confianca(score(q_en, q_musc, q_eq, top[0]))

        # default escolha: 1 pra Forte/Media, vazio pra Fraca (foca atencao)
        prev_e = prev.get(ex["nome"], {})
        escolha = prev_e.get("escolha")
        if escolha in (None, ""):
            escolha = 1 if conf in ("Forte", "Media") else ""

        ws.append([
            ex["nome"], ex["padrao"], ex["musc"], ex["eq"], conf,
            top[0]["name"], top[1]["name"], top[2]["name"],
            escolha, prev_e.get("media_id_final", ""),
            top[0]["id"], top[1]["id"], top[2]["id"],
        ])
        row = ws.max_row
        ws.cell(row, 5).fill = fill_cor[conf]
        # hyperlinks nos candidatos -> imagem 0.jpg
        for col, e in zip((6, 7, 8), top):
            cell = ws.cell(row, col)
            cell.hyperlink = IMG_BASE + e["id"] + "/0.jpg"
            cell.font = link_font

    # dropdown na coluna escolha (I)
    dv = DataValidation(type="list", formula1='"1,2,3,nenhum"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"I2:I{ws.max_row}")

    widths = [26, 18, 22, 16, 10, 34, 34, 34, 9, 30, 1, 1, 1]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    # esconde colunas de id (auxiliares)
    for col in ("K", "L", "M"):
        ws.column_dimensions[col].hidden = True

    wb.save(SAIDA)
    print(f"Gerado: {SAIDA}")


if __name__ == "__main__":
    main()
