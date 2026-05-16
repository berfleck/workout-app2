"""
Sub-H: Remada Landmine split (Annexo 4.1 item 1)

Ação:
  - Renomear 'Remada Landmine' → 'Remada LM Neutra' (mantém pegada=neutra)
  - Inserir nova linha 'Remada LM Aberta' logo após, com pegada=aberta

Ambas mantêm variacao_de='curvada' (família curvada — consistente com
Remada Curvada Barra/Halteres/Smith já no banco).

Nota sobre spec Annexo 4.1 item 1: a spec dizia 'variacao_de = remada curvada'
mas o YAML (ported em Sub-E) já consolidou como 'curvada', alinhado com o
resto da família. Mantemos 'curvada' para consistência com as curvadas existentes.
"""

import openpyxl, copy

XLSX_PATH = "banco_exercicios.xlsx"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    COL = {h: i + 1 for i, h in enumerate(headers) if h}

    # Find Remada Landmine row
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        nome = ws.cell(row=row_idx, column=COL["nome"]).value
        if nome and str(nome).strip() == "Remada Landmine":
            target_row = row_idx
            break

    if target_row is None:
        print("ERROR: 'Remada Landmine' not found in XLSX!")
        return

    print(f"Found 'Remada Landmine' at row {target_row}")

    # Read all values from the original row
    original_values = [ws.cell(row=target_row, column=c).value for c in range(1, len(headers) + 1)]

    # 1. Rename the existing row → 'Remada LM Neutra' (pegada already = 'neutra')
    ws.cell(row=target_row, column=COL["nome"]).value = "Remada LM Neutra"
    print("  Renamed row %d: Remada Landmine -> Remada LM Neutra" % target_row)

    # 2. Insert a new row after target_row for 'Remada LM Aberta'
    ws.insert_rows(target_row + 1)

    new_row = target_row + 1
    for col_idx, val in enumerate(original_values, start=1):
        ws.cell(row=new_row, column=col_idx).value = val

    # Override the differing fields
    ws.cell(row=new_row, column=COL["nome"]).value = "Remada LM Aberta"
    ws.cell(row=new_row, column=COL["pegada"]).value = "aberta"
    ws.cell(row=new_row, column=COL["ativo"]).value = True

    print("  Inserted row %d: 'Remada LM Aberta' (pegada=aberta)" % new_row)

    # Verify
    for r in [target_row, new_row]:
        nome = ws.cell(row=r, column=COL["nome"]).value
        vd   = ws.cell(row=r, column=COL["variacao_de"]).value
        peg  = ws.cell(row=r, column=COL["pegada"]).value
        plan = ws.cell(row=r, column=COL["plano_corporal"]).value
        eq   = ws.cell(row=r, column=COL["equipamento_grupo"]).value
        vp   = ws.cell(row=r, column=COL["variante_pontual"]).value
        print("  row %d: %s | vd=%s | peg=%s | plan=%s | eq=%s | vp=%s" % (
            r, nome, vd, peg, plan, eq, vp
        ))

    wb.save(XLSX_PATH)
    print(f"\nSaved. Total rows: {ws.max_row - 1} exercises")


if __name__ == "__main__":
    main()
